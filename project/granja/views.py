from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.template import loader, RequestContext
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.urls import reverse
from random import randrange
from .models import User, Trabajador, Animal, Suministro, Produccion, Registro, Venta
from django.db.models import Q, Count
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.decorators import login_required
from functools import wraps
from datetime import datetime, date
import json
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings
from django.utils.html import strip_tags
import random
from django.core.serializers.json import DjangoJSONEncoder
from .utils import render_to_pdf, send_invoice_email
import openpyxl
from openpyxl.chart import LineChart, Reference
from itertools import groupby

def index (request):
    ganado = Animal.objects.all().order_by('breed')
    # Filtrar duplicados por raza
    ganado_filtrado = [list(group)[0] for key, group in groupby(ganado, key=lambda x: x.breed)]
    return render(request, 'granja/index.html', {'ganado': ganado_filtrado})

@login_required
def home(request):
    return render(request, 'granja/home.html')

def about (request):
    animales = Animal.objects.all()
    count = 0
    for i in animales:
        count += 1
    return render(request, 'granja/about.html', {'count': count})

def product (request):
    ganado = Animal.objects.all().order_by('breed')
    # Filtrar duplicados por raza
    ganado_filtrado = [list(group)[0] for key, group in groupby(ganado, key=lambda x: x.breed)]
    return render(request, 'granja/product.html', {'ganado': ganado_filtrado})

def team (request):
    return render(request, 'granja/team.html')
"""def home(request):
    return render(request, 'granja/home.html')"""

# Vista para el login
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            # Verifica el rol del user
            trabajador = Trabajador.objects.get(cc=username)
            if trabajador.rol == 'administrador':
                return redirect('administrador')
            elif trabajador.rol == 'trabajador':
                return redirect('worker')
            elif trabajador.rol == 'veterinario':
                return redirect('vet')
        else:
            print('error')
            messages.error(request, 'Nombre de usuario o contraseña incorrectos.')
    else:
        # Verifica si hay una sesión iniciada y redirige a la pagina segun el rol
        if request.user.is_authenticated:
            trabajador = Trabajador.objects.get(cc=request.user.username)
            if trabajador.rol == 'administrador':
                return redirect('administrador')
            elif trabajador.rol == 'trabajador':
                return redirect('worker')
            elif trabajador.rol == 'veterinario':
                return redirect('vet')
    return render(request, 'granja/log_in.html')

# Decorador para restringir el acceso a paginas de otros roles
def rol_required(*roles):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(request, *args, **kwargs):
            if request.user.is_authenticated:
                trabajador = Trabajador.objects.get(cc=request.user.username)
                if trabajador.rol in roles:
                    return view_func(request, *args, **kwargs)
                else:
                    messages.error(request, 'Acceso denegado.')
                    return redirect('index')
            else:
                return redirect('login_view')
        return wrapped_view
    return decorator

# Vista del logout
def log_out(request):
    logout(request)
    return redirect('login')

from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.contrib import messages
from .models import Trabajador

def forgot_password(request):
    if request.method == 'POST':
        cc = request.POST.get('cc')
        email = request.POST.get('email')
        
        try:
            trabajador = Trabajador.objects.get(cc=cc, email=email)
            user = User.objects.get(username=cc)
            
            # Cambiar la contraseña del usuario
            user.set_password(cc)
            user.save()
            
            # Enviar correo electrónico al usuario
            send_mail(
                'Recuperación de contraseña',
                f'Su nueva contraseña es: {cc}',
                'noreply@granja.com',
                [email],
                fail_silently=False,
            )
            
            messages.success(request, 'La contraseña ha sido restablecida. Revise su correo electrónico.')
            return redirect('login')
        
        except Trabajador.DoesNotExist:
            messages.error(request, 'La cédula o el correo electrónico no coinciden con nuestros registros.')
        except User.DoesNotExist:
            messages.error(request, 'No se encontró un usuario con esta cédula.')
    
    return redirect('login')


@login_required
@rol_required('administrador', 'trabajador', 'veterinario')
def change_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        new_password = data.get('new_password')
        
        if new_password:
            user = request.user
            user.set_password(new_password)
            user.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'message': 'La contraseña no puede estar vacía.'})
    return JsonResponse({'success': False, 'message': 'Método no permitido.'})

# ADMINISTRADOR
@login_required
@rol_required('administrador')
def administrador(request):
    trabajador = Trabajador.objects.get(cc=request.user.username)
    context = {
        'trabajador': trabajador,
    }
    
    return render(request, 'granja/administrador.html', context)

# Tabla administrador/animales
@login_required
@rol_required('administrador')
def animals(request):
    user_actual = request.user
    trabajador = Trabajador.objects.get(cc=request.user.username)
    db_data = Animal.objects.all()
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Animal': Animal,
        'user_actual': user_actual,
        'trabajador': trabajador
    }
    return render(request, 'granja/animales.html', context)

@login_required
@rol_required('trabajador')
def animals2(request):
    user_actual = request.user
    db_data = Animal.objects.all()
    trabajador = Trabajador.objects.get(cc=request.user.username)
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Animal': Animal,
        'user_actual': user_actual,
        'trabajador': trabajador
    }
    return render(request, 'granja/animales2.html', context)

# Insertar un animal
@login_required
@rol_required('administrador', 'trabajador')
def insert_animal(request):
    if request.method == 'GET' and 'num_id' in request.GET:
        num_id = request.GET.get('num_id')
        existe = Animal.objects.filter(num_id=num_id).exists()
        return JsonResponse({'existe': existe})
    elif request.method == 'POST':
        try:
            num_id = request.POST['num_identificacion']
            birth_str = request.POST['birthdate']
            birth = datetime.strptime(birth_str, '%Y-%m-%d').date()
            breed = request.POST['breeds']
            color = request.POST['colors']
            sex = request.POST['sexos']
            type = request.POST['tipos']
            if type == 'otro':
                otro_tipo = request.POST['otro_tipo']
            else:
                otro_tipo = 'N/A'
            health = request.POST['health_status']
            if health == 'otro':
                otra_salud = request.POST['otro_estatus']
            else:
                otra_salud = 'N/A'
            aditional = request.POST['aditional_info']
            
            if 'imagen_cow' in request.FILES:
                imagen_cow = request.FILES['imagen_cow']
            else:
                imagen_cow = None
            print(type)
            db_data = Animal(
                num_id=num_id, 
                birth=birth, 
                breed=breed, 
                color=color, 
                sex=sex, 
                type=type,
                otro_tipo=otro_tipo,
                health=health, 
                otra_salud=otra_salud,
                img_cow=imagen_cow, 
                aditional=aditional, 
                encargado_insert=request.user,
                encargado_update=request.user
            )
            db_data.save()
            messages.success(request, 'Registro guardado exitosamente.')
            trabajador = Trabajador.objects.get(cc=request.user.username)
            if trabajador.rol == 'administrador':
                return  HttpResponseRedirect(reverse('animals'))
            elif trabajador.rol == 'trabajador':
                return  HttpResponseRedirect(reverse('animals2'))
        except ValueError as error:
            print(error)
            if trabajador.rol == 'administrador':
                return  HttpResponseRedirect(reverse('animals'))
            elif trabajador.rol == 'trabajador':
                return  HttpResponseRedirect(reverse('animals2'))

def update_edad(request):
    animals = Animal.objects.all()
    for animal in animals:
        actualizar_edad = animal.calcular_edad()
        print(actualizar_edad)
        animal.edad = actualizar_edad
        animal.save()
    return JsonResponse({'mensaje': 'Edades actualizadas correctamente'})
"""def update_edad(request):
    num_id = request.POST.get('num_id')

    try:
        animal = Animal.objects.get(num_id=num_id)
        actualizar_edad = animal.calcular_edad()
        print(actualizar_edad)
        animal.edad = actualizar_edad
        animal.save()
        return JsonResponse({'edad': actualizar_edad})
    except Animal.DoesNotExist:
        print(f'Animal con num_id={num_id} no encontrado.')
        return JsonResponse({'error': 'Animal no encontrado'}, status=404)
    except Exception as e:
        print('Error en la vista update_edad:', str(e))
        return JsonResponse({'error': 'Error interno'}, status=500)"""

# Actualizar un animal
@login_required
@rol_required('administrador', 'trabajador')
def update_animal(request):
    if request.method == 'POST':
        animal_id = request.POST['id']
        num_id = request.POST['num_identificacion']
        birth_str = request.POST['birthdate']
        birth = datetime.strptime(birth_str, '%Y-%m-%d').date()
        breed = request.POST['breeds']
        color = request.POST['colors']
        sex = request.POST['sexos']
        type = request.POST['tipos']
        if type == 'otro':
                otro_tipo = request.POST['otro_tipo']
        else:
            otro_tipo = 'N/A'
        health = request.POST['health_status']
        if health == 'otro':
            otra_salud = request.POST['otro_estatus']
        else:
            otra_salud = 'N/A'
        aditional = request.POST['aditional_info']
        animal = get_object_or_404(Animal, pk=animal_id)

        # Actualización de campos
        animal.num_id = num_id
        animal.birth = birth
        animal.breed = breed
        animal.color = color
        animal.sex = sex
        animal.type = type
        animal.otro_tipo = otro_tipo
        animal.health = health
        animal.otra_salud = otra_salud
        animal.aditional = aditional
        
        # Manejo de la imagen
        if 'imagen_cow' in request.FILES:
            animal.img_cow = request.FILES['imagen_cow']

        animal.save()
        messages.success(request, 'Registro actualizado exitosamente.')
        trabajador = Trabajador.objects.get(cc=request.user.username)
        if trabajador.rol == 'administrador':
            return  HttpResponseRedirect(reverse('animals'))
        elif trabajador.rol == 'trabajador':
            return  HttpResponseRedirect(reverse('animals2'))

# Habilitar el formulario de edición
@login_required
@rol_required('administrador', 'trabajador')
def update_form_animal(request, animal_id):
    db_data = Animal.objects.all()
    db_data_only = Animal.objects.get(pk=animal_id)
    context = {
        "db_data": db_data,
        "update": db_data_only,
        'Animal': Animal,
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/animales.html', context)
    elif trabajador.rol == 'trabajador':
        return render(request, 'granja/animales2.html', context)

# Borrar animal
@login_required
@rol_required('administrador')
def delete_animal(request, animal_id):
    db_data = Animal.objects.filter(id=animal_id)
    db_data.delete()
    messages.success(request, 'El registro ha sido eliminado.')
    return  HttpResponseRedirect(reverse('animals'))
    

# Listar animales con la barra de busqueda
@login_required
@rol_required('administrador', 'trabajador')
def lista_animales(request):
    search_term = request.GET.get('search', '')
    db_data = Animal.objects.all()
    
    if search_term:
        db_data = db_data.filter(
            Q(num_id__icontains=search_term) |
            Q(breed__icontains=search_term) |
            Q(birth__icontains=search_term) |
            Q(type__icontains=search_term)
        )
        """if not db_data.exists():
            messages.error(request, f'No se encontraron registros con el termino ingresado') Falta cuadrar este mensaje"""
    animales = Animal.objects.values_list('num_id', flat=True).distinct()
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Animal': Animal,
        'search_term': search_term,
        'animales': animales
    }
       
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/animales.html', context)
    elif trabajador.rol == 'trabajador':
        return render(request, 'granja/animales2.html', context)

# Tabla administrador/trabajadores
@login_required
@rol_required('administrador')
def workers(request):
    db_data = Trabajador.objects.all()
    trabajador = Trabajador.objects.get(cc=request.user.username)
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Trabajador': Trabajador,
        'trabajador': trabajador
    }
    return render(request, 'granja/trabajadores.html', context)

# Insertar trabajador
@login_required
@rol_required('administrador')
def insert_worker(request):
    if request.method == 'GET' and 'cc' in request.GET:
        cc = request.GET.get('cc')
        existe = Trabajador.objects.filter(cc=cc).exists()
        return JsonResponse({'existe': existe})
    elif request.method == 'POST':
        try:
            name = request.POST['nombre_trabajador']
            cc = request.POST['num_identificacion']
            rol = request.POST['rol']
            direction = request.POST['direccion_residencia']
            tel = request.POST['telefono_contacto']
            email = request.POST['correo_electronico']
            sueldo = request.POST['salario']

            if name == '' or cc == '' or rol == '':
                raise ValueError('Error')
            
            db_data = Trabajador(name = name, cc = cc, rol = rol, direction = direction, tel = tel, email = email,
                                 encargado_insert = request.user, salario = sueldo)
            db_data.save()
            messages.success(request, 'Registro insertado con exito.')
            correo_bienvenida(name, cc, cc, email)

            return HttpResponseRedirect(reverse('workers'))
        except ValueError as error:
            print(error)
            return HttpResponseRedirect(reverse('workers'))

def correo_bienvenida(nombre, username, password, email_trabajador):
    subject = 'Credenciales Trabajador de la Granja "El Manantial"'
    html_content = render_to_string('granja/welcome_email.html',
                                    {'nombre': nombre, 'username': username, 'password': password})
    text_content = strip_tags(html_content)

    email = EmailMultiAlternatives(subject, text_content, settings.DEFAULT_FROM_EMAIL, [email_trabajador])
    email.attach_alternative(html_content, "text/html")
    email.send()

# Actualizar trabajador
@login_required
@rol_required('administrador')
def update_worker(request):
    worker_id = request.POST['id']
    name = request.POST['nombre_trabajador']
    cc = request.POST['num_identificacion']
    rol = request.POST['rol']
    direction = request.POST['direccion_residencia']
    tel = request.POST['telefono_contacto']
    email = request.POST['correo_electronico']
    sueldo = request.POST['salario']
    db_data = Trabajador.objects.get(pk=worker_id)
    db_data.name = name
    db_data.cc = cc
    db_data.rol = rol
    db_data.direction = direction
    db_data.tel = tel
    db_data.email = email
    db_data.salario = sueldo
    db_data.save()
    messages.success(request, 'Registro actualizado con exito.')
    return HttpResponseRedirect(reverse('workers')) 

# Formulario para actualizar trabajador
@login_required
@rol_required('administrador')
def update_form_worker(request, worker_id):
    db_data = Trabajador.objects.all()
    db_data_only = Trabajador.objects.get(pk=worker_id)
    context = {
        "db_data": db_data,
        "update": db_data_only
    }
    
    return render(request, 'granja/trabajadores.html', context)

# Eliminar trabajador
@login_required
@rol_required('administrador')
def delete_worker(request, worker_id):
    trabajador = Trabajador.objects.get(id=worker_id)
    User.objects.filter(username=trabajador.cc).delete()
    Trabajador.objects.filter(id=worker_id).delete()
    messages.success(request, 'El trabajador ha sido eliminado.')
    return HttpResponseRedirect(reverse('workers'))

# Lista de trabajadores
@login_required
@rol_required('administrador')
def lista_trabajadores(request):
    search_term = request.GET.get('search', None)
    db_data = Trabajador.objects.all()
    
    if search_term:
        db_data = db_data.filter(
            Q(name__icontains=search_term) |
            Q(cc__icontains=search_term) |
            Q(rol__icontains=search_term)
        )
    """if not db_data.exists():
            messages.error(request, f'No se encontraron registros con el termino ingresado') Falta cuadrar este mensaje"""
    context = {
        'db_data': db_data,
        'update': None,
        'Trabajador': Trabajador,
        'search_term': search_term
    }
    return render(request, 'granja/trabajadores.html', context)

# Tabla administrador/suministros
@login_required
@rol_required('administrador')
def supplies(request):
    trabajador = Trabajador.objects.get(cc=request.user.username)
    db_data = Suministro.objects.all()
    suministros = Suministro.objects.values_list('name', flat=True).distinct() #Esto es para guardar los nombres de los suministros y llamarlos en los graficos
    animales = Animal.objects.values_list('num_id', flat=True).distinct()
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Suministro': Suministro,
        'suministros': suministros,
        'animales': animales,
        'trabajador': trabajador
    }
    return render(request, 'granja/suministros.html', context)

@login_required
@rol_required('trabajador')
def supplies2(request):
    db_data = Suministro.objects.all()
    suministros = Suministro.objects.values_list('name', flat=True).distinct() #Esto es para guardar los nombres de los suministros y llamarlos en los graficos
    animales = Animal.objects.values_list('num_id', flat=True).distinct()
    trabajador = Trabajador.objects.get(cc=request.user.username)
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Suministro': Suministro,
        'suministros': suministros,
        'animales': animales,
        'trabajador': trabajador
    }
    return render(request, 'granja/suministros2.html', context)

@login_required
@rol_required('administrador', 'trabajador')
def insert_supplie(request):
    if request.method == 'GET' and 'lote' in request.GET:
        lote = request.GET.get('lote')
        existe = Suministro.objects.filter(lote=lote).exists()
        return JsonResponse({'existe': existe})
    elif request.method == 'POST':
        try:
            tipo = request.POST['tipos']
            if tipo == 'otro':
                otro_tipo = request.POST['otro_tipo']
            else:
                otro_tipo = 'N/A'
            nombre = request.POST['name'] 
            cantidad = request.POST['quantity']
            cantidad = float(cantidad)
            medida = request.POST['unit']
            lote = request.POST['lot_number']
            caducidad = request.POST['expiration_date']
            condicion = request.POST['storage_conditions']
            if condicion == 'otro':
                otra_condicion = request.POST['otra_condicion']
            else:
                otra_condicion = 'N/A'
            unidad = request.POST['cost_per_unit']
            unidad = float(unidad)
            total = cantidad * unidad
            descripcion = request.POST['detail']

            db_data = Suministro(tipo = tipo, otro_tipo = otro_tipo, name = nombre, descripcion = descripcion, cantidad = cantidad,
                                medida = medida, lote = lote, caducidad = caducidad, condicion = condicion, otra_condicion = otra_condicion, 
                                unidad = unidad, total = total, encargado_insert = request.user, encargado_update = request.user, 
                                total_actual = total)
                                
            db_data.save()
            messages.success(request, 'Registro insertado con exito.')
            trabajador = Trabajador.objects.get(cc=request.user.username)       
            if trabajador.rol == 'administrador':
                return  HttpResponseRedirect(reverse('supplies'))
            elif trabajador.rol == 'trabajador':
                return  HttpResponseRedirect(reverse('supplies2'))
        except ValueError as error:
            print(error)
            if trabajador.rol == 'administrador':
                return  HttpResponseRedirect(reverse('supplies'))
            elif trabajador.rol == 'trabajador':
                return  HttpResponseRedirect(reverse('supplies2'))

def update_cantidad(request):
    if request.method == 'POST':
        data = json.loads(request.body) # Con esto se llaman los datos del fetch
        suministro_lote = data.get('suministro_lote')
        cantidad_nueva = data.get('cantidad_nueva')
        
        try:
            suministro = Suministro.objects.get(lote=suministro_lote)
            suministro.cantidad = int(cantidad_nueva)
            suministro.save()
            return JsonResponse({'message': 'Cantidad actualizada correctamente.'})
        except Suministro.DoesNotExist:
            return JsonResponse({'error': 'Suministro no encontrado.'}, status=404)
    return JsonResponse({'error': 'Método no permitido.'}, status=405)
"""def update_cantidad(request):
    
    suministro_lote = request.POST.get('suministro_lote')
    cantidad_nueva = request.POST.get('cantidad_nueva')
    print(suministro_lote, cantidad_nueva)
    suministro = Suministro.objects.get(lote=suministro_lote)
    suministro.cantidad = int(cantidad_nueva)
    print(suministro)
    suministro.save()

    return JsonResponse({'message': 'Cantidad actualizada correctamente.'})"""
    
# Actualizar un suministro
@login_required
@rol_required('administrador', 'trabajador')
def update_supplie(request):
    if request.method == 'POST':
        suministro_id = request.POST['id']
        tipo = request.POST['tipos']
        if tipo == 'otro':
                otro_tipo = request.POST['otro_tipo']
        else:
            otro_tipo = 'N/A'
        nombre = request.POST['name']
        cantidad = request.POST['quantity']
        medida = request.POST['unit']
        caducidad = request.POST['expiration_date']
        condicion = request.POST['storage_conditions']
        if condicion == 'otro':
                otra_condicion = request.POST['otra_condicion']
        else:
            otra_condicion = 'N/A'
        descripcion = request.POST['detail']

        if tipo not in Suministro.TIPOS:
            """Suministro.TIPOS.__add__(tuple[tipo])"""
            pass #Quiero que se agregue un nuevo tipo en caso de que no este en la lista
        
        suministro = get_object_or_404(Suministro, pk=suministro_id)

        """if cantidad == 0:
            suministro.estado = 'agotado'
            suministro = Suministro.objects.filter(id=suministro_id)
            suministro.delete()
            messages.info(request, 'El suministro en 0 se elimina')"""
        # Actualización de campos
        unidad = suministro.unidad
        suministro.tipo = tipo
        suministro.otro_tipo = otro_tipo
        suministro.name = nombre
        suministro.cantidad = cantidad
        suministro.medida = medida
        suministro.caducidad = caducidad
        suministro.condicion = condicion
        suministro.otra_condicion = otra_condicion
        suministro.total_actual = float(unidad) * float(cantidad)
        suministro.descripcion = descripcion
        suministro.encargado_update = request.user

        suministro.save()
        messages.success(request, 'Registro actualizado exitosamente.')
        trabajador = Trabajador.objects.get(cc=request.user.username)
        if trabajador.rol == 'administrador':
            return  HttpResponseRedirect(reverse('supplies'))
        elif trabajador.rol == 'trabajador':
            return  HttpResponseRedirect(reverse('supplies2'))

# Habilitar el formulario de edición
@login_required
@rol_required('administrador', 'trabajador')
def update_form_supplie(request, suministro_id):
    db_data = Suministro.objects.all()
    db_data_only = Suministro.objects.get(pk=suministro_id)
    context = {
        "db_data": db_data,
        "update": db_data_only,
        'Suministro': Suministro,
    }
   
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/suministros.html', context)
    elif trabajador.rol == 'trabajador':
        return render(request, 'granja/suministros2.html', context)

# Borrar suministro
@login_required
@rol_required('administrador')
def delete_supplie(request, suministro_id):
    db_data = Suministro.objects.filter(id=suministro_id)
    db_data.delete()
    messages.success(request, 'El registro ha sido eliminado.')
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return  HttpResponseRedirect(reverse('animals'))
    elif trabajador.rol == 'trabajador':
        return  HttpResponseRedirect(reverse('animals2'))

# Listar suministros con la barra de busqueda
@login_required
@rol_required('administrador', 'trabajador')
def lista_suministros(request):
    search_term = request.GET.get('search', None)
    db_data = Suministro.objects.all()
   
    if search_term:
        db_data = db_data.filter(
            Q(lote__icontains=search_term) |
            Q(name__icontains=search_term) |
            Q(tipo__icontains=search_term)
        )
    """if not db_data.exists():
            messages.error(request, f'No se encontraron registros con el termino ingresado') Falta cuadrar este mensaje"""
    suministros = Suministro.objects.values_list('name', flat=True).distinct()
    context = {
        'db_data': db_data[::-1],
        'update': None,
        'Suministro': Suministro,
        'search_term': search_term,
        'suministros': suministros
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/suministros.html', context)
    elif trabajador.rol == 'trabajador':
        return render(request, 'granja/suministros2.html', context)

# Grafico para hacer seguimiento de precios
@login_required
@rol_required('administrador', 'trabajador')
def grafico_precio(request):
    suministro_id = request.GET.get('selectPrecios')
    rango = request.GET.get('rango_fecha')
    fecha_inicio = None
    fecha_fin = None

    if rango:
        fecha_inicio, fecha_fin = rango.split(' - ')

    colors = ['blue', 'orange', 'red', 'black', 'yellow', 'green', 'lightblue', 'purple', 'brown']

    datos = Suministro.objects.filter(
        Q(name=suministro_id) &
        Q(fecha_insert__range=[fecha_inicio, fecha_fin])
    ).order_by('fecha_insert')

    nombre = suministro_id
    fechas = [d.fecha_insert.strftime('%Y-%m-%d') for d in datos]
    precios = [d.unidad for d in datos]
    random_color = colors[randrange(0, len(colors))]

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'fechas_json': json.dumps(fechas, cls=DjangoJSONEncoder),
            'precios_json': json.dumps(precios, cls=DjangoJSONEncoder),
            'nombre_json': json.dumps(nombre, cls=DjangoJSONEncoder),
            'color_json': json.dumps(random_color, cls=DjangoJSONEncoder)
        })

    context = {
        'fechas_json': json.dumps(fechas, cls=DjangoJSONEncoder),
        'precios_json': json.dumps(precios, cls=DjangoJSONEncoder),
        'nombre_json': json.dumps(nombre, cls=DjangoJSONEncoder),
        'color_json': json.dumps(random_color, cls=DjangoJSONEncoder)
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/suministros.html', context)
    elif trabajador.rol == 'trabajador':
        return render(request, 'granja/suministros2.html', context)

# Grafico para verificar existencias
@login_required
@rol_required('administrador', 'trabajador')
def grafico_existencias(request):
    suministro_id = request.GET.get('selectExistencias', None)

    colors = ['blue', 'orange', 'red', 'black', 'yellow', 'green', 'lightblue', 'purple', 'brown']
    
    if not suministro_id:
        return JsonResponse({'error': 'El nombre del suministro es requerido.'}, status=400)
    
    datos = Suministro.objects.filter(
        Q(name=suministro_id)
    ).order_by('lote')

    nombre = suministro_id
    lotes = [d.lote for d in datos]
    existencias = [d.cantidad for d in datos]
    random_color = colors[randrange(0, len(colors))]

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'lotes': lotes,
            'existencias': existencias,
            'nombre': nombre,
            'color': random_color
        })

    context = {
        'lotes_json': json.dumps(lotes, cls=DjangoJSONEncoder),
        'existencias_json': json.dumps(existencias, cls=DjangoJSONEncoder),
        'nombre_json': json.dumps(nombre, cls=DjangoJSONEncoder),
        'color_json': json.dumps(random_color, cls=DjangoJSONEncoder)
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/suministros.html', context)
    elif trabajador.rol == 'trabajador':
        return render(request, 'granja/suministros2.html', context)

"""# Venta
@login_required
@rol_required('administrador', 'trabajador')
def sale(request):
    db_data = Venta.objects.all()
    animal_data = Animal.objects.all()
    producciones = list(Produccion.objects.all().values())

    for ids in db_data:
        if ids.id_vendidas:
            ids.id_vendidas_list = ids.id_vendidas.split(',')

    # Prepara datos JSON para poder ser manejados por JS
    if request.headers.get('accept') == 'application/json':
        animal_data_json = []
        for animal in animal_data:
            animal_dict = {
                'num_id': animal.num_id,
                'animalString': str(animal),
                'edad_categoria': animal.edad_categoria,
                'estado': animal.estado,
            }
            animal_data_json.append(animal_dict)

        data = {
            'animal_data': animal_data_json
        }

        return JsonResponse(data)

    context = {
        'Venta': Venta,
        'db_data': db_data,
        'update': None,
        'animal_data': animal_data,
        'producciones': producciones,
    }

    if animal_data == '':
        messages.error(request, 'No hay animales registrados')
        return render(request, 'granja/ventas.html', context)

    return render(request, 'granja/ventas.html', context)"""

# Insertar venta
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)

@login_required
@rol_required('administrador')
def sale(request):
    db_data = Venta.objects.all()
    animal_data = Animal.objects.all()
    producciones = list(Produccion.objects.all().values())
    trabajador = Trabajador.objects.get(cc=request.user.username)

    for venta in db_data:
        if venta.id_vendidas:
            venta.id_vendidas_list = venta.id_vendidas.split(',')
        else:
            venta.id_vendidas_list = []

        if venta.precio_uni:
            venta.unitario_list = venta.precio_uni.split(',')
        else:
            venta.unitario_list = []
        
        venta.id_unitario_pairs = list(zip(venta.id_vendidas_list, venta.unitario_list))

    if request.headers.get('accept') == 'application/json':
        animal_data_json = []
        for animal in animal_data:
            animal_dict = {
                'num_id': animal.num_id,
                'animalString': str(animal),
                'edad_categoria': animal.edad_categoria,
                'estado': animal.estado,
            }
            animal_data_json.append(animal_dict)

        data = {
            'animal_data': animal_data_json
        }

        return JsonResponse(data)

    context = {
        'Venta': Venta,
        'db_data': db_data,
        'update': None,
        'animal_data': animal_data,
        'producciones': json.dumps(producciones, cls=DateTimeEncoder),
        'trabajador': trabajador,
    }

    return render(request, 'granja/ventas.html', context)
    
@login_required
@rol_required('trabajador')
def sale2(request):
    db_data = Venta.objects.all()
    animal_data = Animal.objects.all()
    producciones = list(Produccion.objects.all().values())
    trabajador = Trabajador.objects.get(cc=request.user.username)

    for venta in db_data:
        if venta.id_vendidas:
            venta.id_vendidas_list = venta.id_vendidas.split(',')
        else:
            venta.id_vendidas_list = []

        if venta.precio_uni:
            venta.unitario_list = venta.precio_uni.split(',')
        else:
            venta.unitario_list = []
        
        venta.id_unitario_pairs = list(zip(venta.id_vendidas_list, venta.unitario_list))

    if request.headers.get('accept') == 'application/json':
        animal_data_json = []
        for animal in animal_data:
            animal_dict = {
                'num_id': animal.num_id,
                'animalString': str(animal),
                'edad_categoria': animal.edad_categoria,
                'estado': animal.estado,
            }
            animal_data_json.append(animal_dict)

        data = {
            'animal_data': animal_data_json
        }

        return JsonResponse(data)

    context = {
        'Venta': Venta,
        'db_data': db_data,
        'update': None,
        'animal_data': animal_data,
        'producciones': json.dumps(producciones, cls=DateTimeEncoder),
        'trabajador': trabajador,
    }

    return render(request, 'granja/ventas2.html', context)

import traceback
@login_required
@rol_required('administrador', 'trabajador')
def insert_sale(request):
    if request.method == 'POST':
        try:
            tipo = request.POST.get('tipoVenta')
            if tipo == 'leche':
                cantidad = float(request.POST.get('cantidadLeche'))
                litro = float(request.POST.get('precioLitro'))
                total = litro * cantidad
                nombre_cliente = request.POST.get('nombreClienteL')
                email_cliente = request.POST.get('emailClienteL')
                adicional = request.POST.get('observacionesLeche')

                producciones = Produccion.objects.order_by('fecha_insert')
                cantidad_restante = cantidad

                for produccion in producciones:
                    if cantidad_restante <= 0:
                        break
                    if produccion.cantidad_leche <= cantidad_restante:
                        cantidad_restante -= produccion.cantidad_leche
                        produccion.cantidad_leche = 0
                        produccion.save()
                    else:
                        produccion.cantidad_leche -= cantidad_restante
                        produccion.save()
                        cantidad_restante = 0

                venta = Venta.objects.create(
                    tipo=tipo,
                    cantidad=cantidad,
                    precio_uni=litro,
                    precio_total=total,
                    nom_cli=nombre_cliente,
                    email_cli=email_cliente,
                    adicional=adicional
                )

                context = {
                    'venta': venta,
                }

                pdf = render_to_pdf('factura.html', context)
                if pdf:
                    filename = f"factura_{venta.id}.pdf"
                    send_invoice_email(venta.email_cli, pdf, filename)

                messages.success(request, 'Venta de leche registrada con éxito.')
                return redirect('sales')
            elif tipo == 'ventaGanado':
                tipo_p = 'Ganado'
                tipo_ganado = request.POST.get('tipoGanado')
                cantidad = request.POST.get('cantidadCabezas')
                ids_vent = request.POST.getlist('idCabezas')
                ids_list = ','.join(ids_vent)

                precio_unitario = request.POST.getlist('precioUnitario')
                unitario_list = ','.join(precio_unitario)
                nombre_cliente = request.POST.get('nombreCliente')
                email_cliente = request.POST.get('emailCliente')
                adicional = request.POST.get('observacionesGanado')

                precio = sum(map(float, precio_unitario))

                venta = Venta.objects.create(
                    tipo=tipo_p,
                    tipo_ganado=tipo_ganado,
                    cantidad=cantidad,
                    id_vendidas=ids_list,
                    precio_total=precio,
                    precio_uni=unitario_list,
                    nom_cli=nombre_cliente,
                    email_cli=email_cliente,
                    adicional=adicional
                )

                for animal_id in ids_vent:
                    animal = Animal.objects.get(num_id=animal_id)
                    animal.estado = 'vendido'
                    animal.save()

                context = {
                    'venta': venta,
                }

                pdf = render_to_pdf('granja/factura.html', context)
                if pdf:
                    filename = f"factura_{venta.id}.pdf"
                    send_invoice_email(venta.email_cli, pdf, filename)

                messages.success(request, 'Venta de ganado registrada con éxito.')
                return redirect('sales')
        except ValueError as error:
            print(f"ValueError: {error}")
            traceback.print_exc()
            messages.error(request, 'Hubo un error al insertar el registro.')
        except Exception as error:
            print(f"Exception: {error}")
            traceback.print_exc()
            messages.error(request, f'Ocurrió un error inesperado: {error}')

    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return HttpResponseRedirect(reverse('sales'))
    if trabajador.rol == 'trabajador':
        return HttpResponseRedirect(reverse('sales2'))


@login_required
@rol_required('administrador', 'trabajador')
def generate_report(request):
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipoReporte')
        rango_fecha = request.POST.get('rango_fecha')
        
        # Parsear el rango de fechas
        fechas = rango_fecha.split(' - ')
        start_date = datetime.strptime(fechas[0], '%Y-%m-%d')
        end_date = datetime.strptime(fechas[1], '%Y-%m-%d')

        # Filtrar los datos en base al rango de fechas y tipo de reporte
        if tipo_reporte == 'leche':
            ventas = Venta.objects.filter(fecha_insert__range=[start_date, end_date], tipo='leche')
        elif tipo_reporte == 'ganado':
            ventas = Venta.objects.filter(fecha_insert__range=[start_date, end_date], tipo='Ganado')
        else:
            ventas = Venta.objects.filter(fecha_insert__range=[start_date, end_date])

        # Crear un libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        # Escribir los encabezados
        headers = ['Fecha', 'Tipo', 'Cantidad', 'Precio Total', 'ID Vendidas', 'Precio Unitario', 'Tipo de Ganado', 'Nombre Cliente', 'Email Cliente', 'Adicional']
        ws.append(headers)

        # Escribir los datos
        for venta in ventas:
            ws.append([
                venta.fecha_insert,
                venta.tipo,
                venta.cantidad,
                venta.precio_total,
                venta.id_vendidas,
                venta.precio_uni,
                venta.tipo_ganado,
                venta.nom_cli,
                venta.email_cli,
                venta.adicional
            ])

        # Ajustar el tamaño de las celdas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Crear un gráfico de líneas
        chart = LineChart()
        chart.title = "Fluctuación de Ventas"
        chart.style = 10
        chart.y_axis.title = 'Precio Total'
        chart.x_axis.title = 'Fecha'

        # Referenciar los datos para el gráfico
        data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Añadir el gráfico a la hoja
        ws.add_chart(chart, "K5")

        # Guardar el archivo Excel en un objeto HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Reporte_Ventas_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        wb.save(response)

        return response
    else:
        trabajador = Trabajador.objects.get(cc=request.user.username)
        if trabajador.rol == 'administrador':
            return render(request, 'granja/ventas.html')
        if trabajador.rol == 'trabajador':
            return render(request, 'granja/ventas2.html')

# Borrar venta
@login_required
@rol_required('administrador')
def delete_sale(request, venta_id):
    venta = Venta.objects.get(id=venta_id)
    venta.delete()
    messages.success(request, 'Venta eliminada con éxito.')
    
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return HttpResponseRedirect(reverse('sales'))
    if trabajador.rol == 'trabajador':
        return HttpResponseRedirect(reverse('sales2'))

@login_required
@rol_required('administrador', 'trabajador')
def lista_sale(request):
    search_term = request.GET.get('search', None)
    db_data = Venta.objects.all()
    if search_term:
        db_data = db_data.filter(
            Q(fecha_insert__icontains=search_term) |
            Q(id_animal__icontains=search_term) |
            Q(tipo__icontains=search_term) |
            Q(tipo_ganado__icontains=search_term)
        )
   
    registros = Venta.objects.values_list('tipo', flat=True).distinct()
    context = {
        'registros': registros,
        'db_data': db_data[::-1],
        'search_term': search_term,
        'Venta': Venta
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/ventas.html', context)
    if trabajador.rol == 'trabajador':
        return render(request, 'granja/ventas2.html', context)

# Tabla administrador/produccion
@login_required
@rol_required('administrador')
def production(request):
    db_data = Produccion.objects.all()
    animal_data = Animal.objects.all()
    trabajador = Trabajador.objects.get(cc=request.user.username)

    # Prepara datos JSON para poder ser manejados por JS
    if request.headers.get('accept') == 'application/json':
        animal_data_json = []
        for animal in animal_data:
            animal_dict = {
                'num_id': animal.num_id,
                'animalString': str(animal),
                'edad_categoria': animal.edad_categoria,
                'estado': animal.estado,
            }
            animal_data_json.append(animal_dict)

        data = {
            'animal_data': animal_data_json
        }

        return JsonResponse(data)

    context = {
        'Produccion': Produccion,
        'db_data': db_data,
        'update': None,
        'animal_data': animal_data,
        'trabajador': trabajador
    }

    if animal_data == '':
        messages.error(request, 'No hay animales registrados')
    return render(request, 'granja/produccion.html', context)
            
        
@login_required
@rol_required('trabajador')
def production2(request):
    db_data = Produccion.objects.all()
    animal_data = Animal.objects.all()
    trabajador = Trabajador.objects.get(cc=request.user.username)

    # Prepara datos JSON para poder ser manejados por JS
    if request.headers.get('accept') == 'application/json':
        animal_data_json = []
        for animal in animal_data:
            animal_dict = {
                'num_id': animal.num_id,
                'animalString': str(animal),
                'edad_categoria': animal.edad_categoria,
                'estado': animal.estado,
            }
            animal_data_json.append(animal_dict)

        data = {
            'animal_data': animal_data_json
        }

        return JsonResponse(data)

    context = {
        'Produccion': Produccion,
        'db_data': db_data,
        'update': None,
        'animal_data': animal_data,
        'trabajador': trabajador
    }

    if animal_data == '':
        messages.error(request, 'No hay animales registrados')
        trabajador = Trabajador.objects.get(cc=request.user.username)
        if trabajador.rol == 'administrador':
            return render(request, 'granja/produccion.html', context)
        if trabajador.rol == 'trabajador':
            return render(request, 'granja/produccion2.html', context)

    return render(request, 'granja/produccion2.html', context)

# Insertar produccion
@login_required
@rol_required('administrador', 'trabajador')
def insert_production(request):
    if request.method == 'POST':
        cantidad = request.POST.get('cantidadLeche')
        vaca = request.POST.get('idVaca')
        adicional = request.POST.get('observacionesLeche')
        produccion = Produccion(cantidad_leche=cantidad, vaca=vaca, adicional=adicional)
        produccion.save()
        # Esto lo necesito para la venta de ganado
        """elif tipo == 'ventaGanado':
            tipo_ganado = request.POST.get('tipoGanado')
            cantidad = request.POST.get('cantidadCabezas')
            ids_vent = request.POST.getlist('idCabezas')
            ids_list = ','.join(ids_vent) 
            precio = request.POST.get('precioVenta')
            adicional = request.POST.get('observacionesGanado')
            produccion = Produccion(tipo=tipo, tipo_ganado=tipo_ganado, cantidad_vendidas=cantidad, id_vendidas=ids_list, precio=precio, adicional=adicional,
                                cantidad_leche=0, vaca=0)
            produccion.save()

            for animal_id in ids_vent:
                animal = Animal.objects.get(num_id=animal_id)
                animal.estado = 'vendido'
                animal.save()"""
        messages.success(request, 'Registro insertado con éxito.')
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return HttpResponseRedirect(reverse('production'))
    if trabajador.rol == 'trabajador':
        return HttpResponseRedirect(reverse('production2'))

#Eliminar producción
@login_required
@rol_required('administrador')
def delete_production(request, produccion_id):
    db_data = Produccion.objects.filter(id=produccion_id)
    db_data.delete()
    messages.success(request, 'El registro ha sido eliminado.')
    return HttpResponseRedirect(reverse('production'))

#Listar producción
@login_required
@rol_required('administrador', 'trabajor')
def lista_produccion(request):
    search_term = request.GET.get('search', None)
    db_data = Produccion.objects.all()
    if search_term:
        db_data = db_data.filter(
            Q(vaca__icontains=search_term) |
            Q(fecha_insert__icontains=search_term)
        )
    """if not db_data.exists():
            messages.error(request, f'No se encontraron registros con el termino ingresado') Falta cuadrar este mensaje"""
    producciones = Produccion.objects.values_list('tipo', flat=True).distinct()
    context = {
        'producciones': producciones,
        'db_data': db_data[::-1],
        'search_term': search_term,
        'Produccion': Produccion
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/produccion.html', context)
    if trabajador.rol == 'trabajador':
        return render(request, 'granja/produccion2.html', context)

# Tabla administrador/registros-salud
@login_required
@rol_required('administrador')
def record(request):
    db_data = Registro.objects.all()
    animal_data = Animal.objects.all()
    toros_list = Animal.objects.filter(edad_categoria='toro')
    vacas_list = Animal.objects.filter(edad_categoria='vaca')
    vacunas_list = Suministro.objects.filter(tipo='vacuna')
    medicamento_list = Suministro.objects.filter(tipo='medicamento')
    trabajador = Trabajador.objects.get(cc=request.user.username)
    context = {
        'Registro': Registro,
        'db_data': db_data[::-1],
        'animal_data': animal_data,
        'toros_list': toros_list,
        'vacunas_list': vacunas_list,
        'medicamento_list': medicamento_list,
        'vacas_list': vacas_list,
        'trabajador': trabajador
    }
    return render(request, 'granja/registros_vet.html', context)

@login_required
@rol_required('veterinario')
def record2(request):
    db_data = Registro.objects.all()
    animal_data = Animal.objects.all()
    toros_list = Animal.objects.filter(edad_categoria='toro')
    vacas_list = Animal.objects.filter(edad_categoria='vaca')
    vacunas_list = Suministro.objects.filter(tipo='vacuna')
    medicamento_list = Suministro.objects.filter(tipo='medicamento')
    trabajador = Trabajador.objects.get(cc=request.user.username)
    context = {
        'Registro': Registro,
        'db_data': db_data[::-1],
        'animal_data': animal_data,
        'toros_list': toros_list,
        'vacunas_list': vacunas_list,
        'medicamento_list': medicamento_list,
        'vacas_list': vacas_list,
        'trabajador': trabajador
    }
    return render(request, 'granja/registros_vet2.html', context)

# Insert registro de salud
@login_required
@rol_required('administrador', 'veterinario')
def insert_record(request):
    if request.method == 'POST':
        tipo = request.POST.get('registro')
        animal_id = request.POST.get('animal_id')
        animal = Animal.objects.get(num_id=animal_id)

        if tipo == 'estado_general':
            sintomas = request.POST.get('sintomas')
            descripcion = request.POST.get('descripcion_registro')
            registro = Registro(
                tipo=tipo,
                id_animal=animal_id,
                general=sintomas,
                descripcion=descripcion,
                encargado_insert=request.user,
            )
            registro.save()
            if sintomas == 'si':
                animal.health = 'enfermo'
                animal.encargado_update = request.user
            else:
                animal.health = animal.health
            animal.save()
        elif tipo == 'registro_vacunacion':
            nom_vacuna = request.POST.get('nombre_vacuna')
            descripcion = request.POST.get('descripcion_registro')
            vacuna = Suministro.objects.get(name=nom_vacuna)
            registro = Registro(
                tipo=tipo,
                id_animal=animal,
                nom_vacuna=vacuna,
                descripcion=descripcion,
                encargado_insert=request.user
            )
            registro.save()
            vacuna.cantidad -= 1
            vacuna.encargado_update = request.user
            vacuna.save()

        elif tipo == 'registro_tratamiento':
            nom_medicamento = request.POST.get('nombre_medicamento')
            razon = request.POST.get('razon_tratamiento')
            dosis = request.POST.get('dosis_administrada')
            medicamento = Suministro.objects.get(name=nom_medicamento)
            registro = Registro(
                tipo=tipo,
                id_animal=animal,
                nom_medicamento=medicamento,
                razon=razon,
                dosis=dosis,
                encargado_insert=request.user
            )
            registro.save()
            medicamento.cantidad -= 1
            medicamento.encargado_update = request.user
            medicamento.save()
        elif tipo == 'registro_peso':
            peso = request.POST.get('peso_actual')
            registro = Registro(
                tipo=tipo,
                id_animal=animal,
                peso=peso,
                encargado_insert=request.user
                )
            registro.save()
            animal.peso = peso
            animal.encargado_update = request.user
            animal.save()
        elif tipo == 'registro_inseminaciones':
            #fecha = request.POST.get('fecha_inseminacion)
            metodo = request.POST.get('metodo_inseminacion')
            id_toro = request.POST.get('id_toro')
            toro = Animal.objects.get(num_id=id_toro)
            registro = Registro(
                tipo=tipo,
                id_animal=animal,
                tipo_inseminacion=metodo,
                mat_genetico=toro,
                encargado_insert=request.user
            )
            registro.save()
            if animal.health != 'otro':
                animal.health = 'otro'
                animal.otra_salud = 'en gestación'
                animal.encargado_update = request.user
                animal.save()
            elif animal.health == 'otro':
                animal.otra_salud = 'en gestación'
                animal.encargado_update = request.user
                animal.save()
        elif tipo == 'registro_partos':
            numero_crias = request.POST.get('numero_crias')
            estado_crias = request.POST.get('estado_crias')
            registro = Registro(
                tipo=tipo,
                id_animal=animal,
                num_crias=numero_crias,
                estado=estado_crias,
                encargado_insert=request.user
            )
            registro.save()
            if animal.health != 'otro':
                animal.health = 'otro'
                animal.otra_salud = 'en lactancia'
                animal.encargado_update = request.user
                animal.save()
            elif animal.health == 'otro':
                animal.otra_salud = 'en lactancia'
                animal.encargado_update = request.user
                animal.save()
        messages.success(request, 'Registro insertado con éxito.')
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return HttpResponseRedirect(reverse('record'))
    if trabajador.rol == 'veterinario':
        return HttpResponseRedirect(reverse('record2'))

@login_required
@rol_required('administrador')
def delete_record(request, record_id):
    db_data = Registro.objects.filter(id=record_id)
    db_data.delete()
    messages.success(request, 'El registro ha sido eliminado.')
    return HttpResponseRedirect(reverse('record'))

@login_required
@rol_required('administrador', 'veterinario')
def lista_record(request):
    search_term = request.GET.get('search', None)
    db_data = Registro.objects.all()
    if search_term:
        db_data = db_data.filter(
            Q(fecha_insert__icontains=search_term) |
            Q(id_animal__icontains=search_term) |
            Q(tipo__icontains=search_term)
        )
   
    registros = Registro.objects.values_list('tipo', flat=True).distinct()
    context = {
        'registros': registros,
        'db_data': db_data[::-1],
        'search_term': search_term,
        'Registro': Registro
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/registros_vet.html', context)
    if trabajador.rol == 'veterinario':
        return render(request, 'granja/registros_vet2.html', context)

@login_required
@rol_required('administrador', 'veterinario')
def grafico_salud(request):
    tipo_registro = request.GET.get('tipo_registro')
    rango = request.GET.get('rango_fecha')
    fecha_inicio = None
    fecha_fin = None

    if rango:
        fecha_inicio, fecha_fin = rango.split(' - ')

    if not (tipo_registro and fecha_inicio and fecha_fin):
        return JsonResponse({'error': 'Todos los campos son requeridos.'}, status=400)

    registros = Registro.objects.filter(
        tipo=tipo_registro,
        fecha_insert__range=[fecha_inicio, fecha_fin]
    ).values('fecha_insert').annotate(cantidad=Count('id')).order_by('fecha_insert')

    fechas = [registro['fecha_insert'].strftime('%Y-%m-%d') for registro in registros]
    cantidades = [registro['cantidad'] for registro in registros]
    nombre = tipo_registro.replace('_', ' ').title()
    colors = ['blue', 'orange', 'red', 'black', 'yellow', 'green', 'lightblue', 'purple', 'brown']
    random_color = random.choice(colors)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'fechas': fechas,
            'cantidades': cantidades,
            'nombre': nombre,
            'color': random_color
        })

    context = {
        'fechas_json': json.dumps(fechas, cls=DjangoJSONEncoder),
        'cantidades_json': json.dumps(cantidades, cls=DjangoJSONEncoder),
        'nombre_json': json.dumps(nombre, cls=DjangoJSONEncoder),
        'color_json': json.dumps(random_color, cls=DjangoJSONEncoder)
    }
    trabajador = Trabajador.objects.get(cc=request.user.username)
    if trabajador.rol == 'administrador':
        return render(request, 'granja/registros_vet.html', context)
    if trabajador.rol == 'veterinario':
        return render(request, 'granja/registros_vet2.html', context)

@login_required
@rol_required('administrador', 'veterinario')
def grafico_crias(request):
    animal_id = request.GET.get('listaVacas')
    rango_fecha = request.GET.get('rango_fecha_crias')
    fecha_inicio, fecha_fin = rango_fecha.split(' - ')
    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')

    registros = Registro.objects.filter(id_animal=animal_id, fecha_insert__range=[fecha_inicio, fecha_fin], tipo='registro_partos')
    
    fechas = [registro.fecha_insert.strftime('%Y-%m-%d') for registro in registros]
    cantidades = [int(registro.num_crias) for registro in registros]

    data = {
        'fechas': fechas,
        'cantidades': cantidades,
        'nombre': get_object_or_404(Animal, num_id=animal_id).num_id,
        'color': '#ff7f50'
    }

    return JsonResponse(data)

# TRABAJADOR
@login_required
@rol_required('trabajador')
def worker(request):
    trabajador = Trabajador.objects.get(cc=request.user.username)

    context = {
        'trabajador': trabajador
    }
    return render(request, 'granja/trabajador.html', context)

"""# Tabla trabajador/animales
@login_required
@rol_required('trabajador')
def animals2(request):
    return render(request, 'granja/animales2.html')

# Tabla trabajador/produccion
@login_required
@rol_required('trabajador')
def production2(request):
    return render(request, 'granja/produccion2.html')

# Tabla trabajador/suministros
@login_required
@rol_required('trabajador')
def supplies2(request):
    return render(request, 'granja/suministros2.html')"""

# VETERINARIO
@login_required
@rol_required('veterinario')
def vet(request):
    trabajador = Trabajador.objects.get(cc=request.user.username)

    context = {
        'trabajador': trabajador
    }
    return render(request, 'granja/veterinario.html', context)

"""# Tabla veterinario/registros-salud
@login_required
@rol_required('veterinario')
def record2(request):
    return render(request, 'granja/registros_vet2.html')"""